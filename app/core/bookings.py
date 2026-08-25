"""Excel booking store.

openpyxl rewrites the whole workbook on save, so two concurrent bookings would
clobber each other. Every write takes a process-wide lock, saves to a `.tmp.xlsx`
and then `os.replace()`s it into place (atomic on the same filesystem).

If the salon owner has bookings.xlsx open in Excel the replace raises OSError on
Windows; we catch it and append the row to a CSV so a booking is never lost.
"""

from __future__ import annotations

import csv
import logging
import os
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import config
from app.adapters import sheets

log = logging.getLogger(__name__)

# Shared with the Google Sheet mirror so the two never drift apart.
HEADERS: list[str] = config.BOOKING_HEADERS
COLUMN_WIDTHS = [12, 20, 22, 20, 18, 18, 16, 12, 28]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # dark blue
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=11)

_lock = threading.Lock()


def _new_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"
    ws.append(HEADERS)
    for idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = COLUMN_WIDTHS[idx - 1]
    ws.freeze_panes = "A2"
    return wb


def _load() -> Workbook:
    if config.BOOKINGS_XLSX.exists():
        try:
            return load_workbook(config.BOOKINGS_XLSX)
        except Exception:  # noqa: BLE001 - corrupt file must not stop bookings
            log.exception("bookings.xlsx unreadable, starting a fresh workbook")
    return _new_workbook()


def _id_number(value: Any) -> int:
    if isinstance(value, str) and value.startswith("BK-"):
        try:
            return int(value.split("-", 1)[1])
        except ValueError:
            return 0
    return 0


def _next_booking_id(ws: Any) -> str:
    """Highest id across the sheet AND the CSV fallback, so refs never repeat."""
    highest = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        highest = max(highest, _id_number(row[0]))
    if config.BOOKINGS_CSV_FALLBACK.exists():
        try:
            with config.BOOKINGS_CSV_FALLBACK.open(newline="", encoding="utf-8") as handle:
                for record in csv.reader(handle):
                    if record:
                        highest = max(highest, _id_number(record[0]))
        except OSError:
            log.exception("Could not read CSV fallback for id sequencing")
    return f"BK-{highest + 1:04d}"


def _csv_fallback(row: list[Any]) -> None:
    path: Path = config.BOOKINGS_CSV_FALLBACK
    exists = path.exists()
    with path.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        if not exists:
            writer.writerow(HEADERS)
        writer.writerow(row)
    log.warning("Booking written to CSV fallback %s", path)


def save_booking(
    *,
    customer_name: str,
    whatsapp_number: str,
    service: str,
    date: str,
    time: str,
    channel: str,
    status: str = "Pending",
    notes: str = "",
) -> str:
    """Append one booking row. Returns the booking reference (e.g. BK-0001)."""
    received_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    note_text = notes or f"via {channel}"

    row: list[Any] = []
    with _lock:
        wb = _load()
        ws = wb.active
        booking_id = _next_booking_id(ws)
        row = [
            booking_id,
            received_at,
            customer_name,
            whatsapp_number,
            service,
            date,
            time,
            status,
            note_text,
        ]
        ws.append(row)
        last = ws.max_row
        for idx in range(1, len(HEADERS) + 1):
            ws.cell(row=last, column=idx).font = BODY_FONT

        tmp = config.BOOKINGS_XLSX.with_suffix(".tmp.xlsx")
        try:
            wb.save(tmp)
            os.replace(tmp, config.BOOKINGS_XLSX)
        except OSError as exc:
            # Typically: the owner has the file open in Excel and it is locked.
            log.error("Could not write bookings.xlsx (%s) — using CSV fallback", exc)
            try:
                if tmp.exists():
                    tmp.unlink()
            except OSError:
                pass
            _csv_fallback(row)
        finally:
            wb.close()

    log.info("Booking %s saved (%s, %s %s, %s)", booking_id, service, date, time, channel)

    # Mirror to the owner's Google Sheet on a daemon thread. Excel above is the source
    # of truth; a Sheets outage must not delay or break the customer's confirmation.
    try:
        sheets.append_booking_async(dict(zip(HEADERS, row)))
    except Exception:  # noqa: BLE001 - even failing to start the thread is non-fatal
        log.exception("Could not dispatch booking %s to Google Sheets", booking_id)

    return booking_id


def list_bookings() -> list[dict[str, Any]]:
    """Read the sheet back, plus anything that landed in the CSV fallback."""
    rows: list[dict[str, Any]] = []
    with _lock:
        if config.BOOKINGS_XLSX.exists():
            try:
                wb = load_workbook(config.BOOKINGS_XLSX, read_only=True)
                ws = wb.active
                for values in ws.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True):
                    if values and any(v is not None for v in values):
                        rows.append({h: ("" if v is None else str(v)) for h, v in zip(HEADERS, values)})
                wb.close()
            except Exception:  # noqa: BLE001
                log.exception("Failed reading bookings.xlsx")
        if config.BOOKINGS_CSV_FALLBACK.exists():
            try:
                with config.BOOKINGS_CSV_FALLBACK.open(newline="", encoding="utf-8") as handle:
                    for record in csv.DictReader(handle):
                        rows.append({h: record.get(h, "") for h in HEADERS})
            except OSError:
                log.exception("Failed reading CSV fallback")
    return rows
